import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import re
import io
import os
import textwrap
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as ExcelImage
import openai

# OpenAI klientas
client = openai.OpenAI(api_key=st.secrets["openai_api_key"])

st.set_page_config(page_title="Klaidų analizės dashboard", layout="wide")

# ----------------------------
# STILIUS
# ----------------------------
st.markdown("""
<style>
.metric-card {
    padding: 18px 20px;
    border-radius: 16px;
    color: white;
    box-shadow: 0 6px 18px rgba(0,0,0,0.08);
    margin-bottom: 10px;
}
.metric-title {
    font-size: 14px;
    opacity: 0.9;
    margin-bottom: 8px;
}
.metric-value {
    font-size: 30px;
    font-weight: 700;
    line-height: 1.1;
}
.metric-green { background: linear-gradient(135deg, #0f9d58, #34a853); }
.metric-red { background: linear-gradient(135deg, #d93025, #ea4335); }
.metric-blue { background: linear-gradient(135deg, #1a73e8, #4285f4); }
.metric-purple { background: linear-gradient(135deg, #7b1fa2, #9c27b0); }

.block-container {
    padding-top: 1.2rem;
    padding-bottom: 2rem;
}

div[data-testid="stDataFrame"] {
    border-radius: 12px;
    overflow: hidden;
}
</style>
""", unsafe_allow_html=True)

st.title("📊 Klaidų analizės dashboard")
st.caption("Analizėje naudojama: **Klaidos priežastis** iš **O** stulpelio ir **Klaidos** iš **P** stulpelio.")

uploaded_file = st.file_uploader("📎 Pasirinkite Excel failą", type=["xlsx"])


# ----------------------------
# PAGALBINĖS FUNKCIJOS
# ----------------------------
def extract_month(text):
    if isinstance(text, str):
        match = re.search(
            r"\b(KOVAS|VASARIS|SAUSIS|BALANDIS|GEGUŽĖ|BIRŽELIS|LIEPA|RUGPJŪTIS|RUGSĖJIS|SPALIS|LAPKRITIS|GRUODIS)\b",
            text.upper()
        )
        if match:
            return match.group(1).capitalize()
    return "Nežinoma"


def clean_text(value):
    if pd.isna(value):
        return None
    text = str(value).strip()
    return text if text else None


def generate_insight(row):
    klaidos = row["Su_klaidomis"]
    procentas = row["Klaidų_procentas"]
    saskaitu = row["Sąskaitų_skaičius"]
    menuo = row["Mėnuo"]

    if klaidos == 0:
        return f"✅ {menuo}: jokių klaidų – puikus rezultatas."
    elif saskaitu < 15 and procentas >= 15:
        return f"⚠️ {menuo}: nors klaidų tik {klaidos}, jos sudaro {procentas:.2f}% – mažas kiekis padidina procentinę įtaką."
    elif procentas >= 20:
        return f"🔴 {menuo}: labai aukštas klaidų procentas ({procentas:.2f}%) – būtina peržiūrėti procesą."
    elif procentas >= 15:
        return f"🟠 {menuo}: padidėjęs klaidų procentas ({procentas:.2f}%) – verta ieškoti priežasčių."
    else:
        return f"🟢 {menuo}: klaidų lygis ({procentas:.2f}%) yra kontroliuojamas."


def safe_add_image(ws, image_path, anchor):
    if os.path.exists(image_path):
        img = ExcelImage(image_path)
        img.anchor = anchor
        ws.add_image(img)


def wrap_label(text, width=18):
    if pd.isna(text):
        return "Nenurodyta"
    return "\n".join(textwrap.wrap(str(text), width=width))


def ellipsis_label(text, max_len=30):
    if pd.isna(text):
        return "Nenurodyta"
    text = str(text)
    return text if len(text) <= max_len else text[:max_len - 3] + "..."


if uploaded_file:
    df = pd.read_excel(uploaded_file)

    # ----------------------------
    # PRIVALOMŲ STULPELIŲ PATIKRA
    # ----------------------------
    required_named_columns = ["Klientas", "Užsakovas", "Sąskaitos faktūros Nr.", "Siuntėjas"]
    missing_named = [col for col in required_named_columns if col not in df.columns]

    if missing_named:
        st.error(f"Faile trūksta šių stulpelių: {', '.join(missing_named)}")
        st.stop()

    if df.shape[1] < 16:
        st.error("Faile nepakanka stulpelių. Reikia bent 16 stulpelių, kad būtų galima paimti O ir P.")
        st.stop()

    # ----------------------------
    # DUOMENŲ PARUOŠIMAS
    # O = 14 indeksas, P = 15 indeksas
    # ----------------------------
    df["Klaidos_priežastis"] = df.iloc[:, 14].apply(clean_text)
    df["Klaidos"] = df.iloc[:, 15].apply(clean_text)
    df["Mėnuo"] = df["Klientas"].apply(extract_month)
    df["Yra klaida"] = df["Klaidos"].notna()

    menesiu_tvarka = [
        "Sausis", "Vasaris", "Kovas", "Balandis", "Gegužė", "Birželis",
        "Liepa", "Rugpjūtis", "Rugsėjis", "Spalis", "Lapkritis", "Gruodis"
    ]

    visi_menesiai = sorted(
        df["Mėnuo"].dropna().unique(),
        key=lambda x: menesiu_tvarka.index(x) if x in menesiu_tvarka else 99
    )
    visi_siuntejai = sorted(df["Siuntėjas"].dropna().astype(str).unique().tolist())
    visi_uzsakovai = sorted(df["Užsakovas"].dropna().astype(str).unique().tolist())

    # ----------------------------
    # FILTRAI
    # ----------------------------
    st.sidebar.header("🎛️ Filtrai")

    pasirinkti_menesiai = st.sidebar.multiselect(
        "📆 Mėnesiai",
        visi_menesiai,
        default=visi_menesiai
    )

    pasirinkti_siuntejai = st.sidebar.multiselect(
        "📨 Siuntėjai",
        visi_siuntejai,
        default=visi_siuntejai
    )

    pasirinkti_uzsakovai = st.sidebar.multiselect(
        "🏢 Užsakovai",
        visi_uzsakovai,
        default=visi_uzsakovai
    )

    rodyti_tik_klaidas = st.sidebar.checkbox("Rodyti tik įrašus su klaidomis", value=False)

    df_filtered = df[
        (df["Mėnuo"].isin(pasirinkti_menesiai)) &
        (df["Siuntėjas"].astype(str).isin(pasirinkti_siuntejai)) &
        (df["Užsakovas"].astype(str).isin(pasirinkti_uzsakovai))
    ].copy()

    if rodyti_tik_klaidas:
        df_filtered = df_filtered[df_filtered["Yra klaida"] == True].copy()

    if df_filtered.empty:
        st.warning("Pagal pasirinktus filtrus duomenų nerasta.")
        st.stop()

    # ----------------------------
    # SUVESTINĖ PAGAL MĖNESIUS
    # ----------------------------
    summary = df_filtered.groupby("Mėnuo").agg(
        Sąskaitų_skaičius=("Sąskaitos faktūros Nr.", "nunique"),
        Su_klaidomis=("Yra klaida", "sum")
    ).reset_index()

    summary["Klaidų_procentas"] = (
        summary["Su_klaidomis"] / summary["Sąskaitų_skaičius"] * 100
    ).round(2)

    summary["Mėnesio_nr"] = summary["Mėnuo"].apply(
        lambda x: menesiu_tvarka.index(x) if x in menesiu_tvarka else -1
    )
    summary = summary.sort_values("Mėnesio_nr").drop(columns="Mėnesio_nr")

    max_skaicius = summary["Sąskaitų_skaičius"].max() if not summary.empty else 1
    summary["Sąskaitų_procentas"] = (
        summary["Sąskaitų_skaičius"] / max_skaicius * 100
    ).round(2)

    summary["Įžvalga"] = summary.apply(generate_insight, axis=1)

    # ----------------------------
    # KLAIDŲ SĄRAŠAS
    # ----------------------------
    klaidos = df_filtered[df_filtered["Yra klaida"] == True][[
        "Mėnuo",
        "Užsakovas",
        "Sąskaitos faktūros Nr.",
        "Klaidos_priežastis",
        "Klaidos",
        "Siuntėjas"
    ]].copy()

    # ----------------------------
    # KPI
    # ----------------------------
    viso_dokumentu = int(df_filtered["Sąskaitos faktūros Nr."].nunique())
    viso_klaidu = int(df_filtered["Yra klaida"].sum())
    klaidu_proc = round((viso_klaidu / viso_dokumentu * 100), 2) if viso_dokumentu else 0.0
    be_klaidu = viso_dokumentu - viso_klaidu

    k1, k2, k3, k4 = st.columns(4)

    with k1:
        st.markdown(f"""
        <div class="metric-card metric-blue">
            <div class="metric-title">Viso sąskaitų</div>
            <div class="metric-value">{viso_dokumentu}</div>
        </div>
        """, unsafe_allow_html=True)

    with k2:
        st.markdown(f"""
        <div class="metric-card metric-red">
            <div class="metric-title">Su klaidomis</div>
            <div class="metric-value">{viso_klaidu}</div>
        </div>
        """, unsafe_allow_html=True)

    with k3:
        st.markdown(f"""
        <div class="metric-card metric-purple">
            <div class="metric-title">Klaidų procentas</div>
            <div class="metric-value">{klaidu_proc}%</div>
        </div>
        """, unsafe_allow_html=True)

    with k4:
        st.markdown(f"""
        <div class="metric-card metric-green">
            <div class="metric-title">Be klaidų</div>
            <div class="metric-value">{be_klaidu}</div>
        </div>
        """, unsafe_allow_html=True)

    # ----------------------------
    # 1 EILĖ: MĖNESIŲ SUVESTINĖ + GRAFIKAS
    # ----------------------------
    left, right = st.columns([1.15, 1])

    with left:
        st.subheader("📋 Suvestinė pagal mėnesius")
        st.dataframe(summary, use_container_width=True, height=420)

    with right:
        st.subheader("📈 Normalizuotas palyginimas")
        fig_months, ax_months = plt.subplots(figsize=(9, 5.5))
        ax_months.plot(
            summary["Mėnuo"],
            summary["Sąskaitų_procentas"],
            label="Sąskaitų kiekis (%)",
            marker="o",
            linewidth=2
        )
        ax_months.plot(
            summary["Mėnuo"],
            summary["Klaidų_procentas"],
            label="Klaidų procentas (%)",
            marker="o",
            linewidth=2
        )
        ax_months.set_ylabel("Procentai (%)")
        ax_months.set_xlabel("Mėnuo")
        ax_months.set_ylim(0, 100)
        ax_months.legend()
        ax_months.grid(True, alpha=0.3)
        plt.title("Sąskaitų kiekis ir klaidų procentas")
        st.pyplot(fig_months)

    st.subheader("🔎 Įžvalgos pagal mėnesius")
    st.dataframe(
        summary[[
            "Mėnuo",
            "Klaidų_procentas",
            "Sąskaitų_skaičius",
            "Sąskaitų_procentas",
            "Su_klaidomis",
            "Įžvalga"
        ]],
        use_container_width=True
    )

    # ----------------------------
    # ANALIZĖ, JEI YRA KLAIDŲ
    # ----------------------------
    if not klaidos.empty:
        # Klaidos pagal priežastį
        priezastys = klaidos["Klaidos_priežastis"].fillna("Nenurodyta").value_counts().reset_index()
        priezastys.columns = ["Klaidos priežastis", "Klaidų skaičius"]

        # Klaidos pagal siuntėją
        siuntejai_klaidos = klaidos["Siuntėjas"].fillna("Nenurodyta").value_counts().reset_index()
        siuntejai_klaidos.columns = ["Siuntėjas", "Klaidų skaičius"]

        # Visi dokumentai pagal siuntėją
        siuntejai_visi = df_filtered.groupby("Siuntėjas").agg(
            Dokumentų_skaičius=("Sąskaitos faktūros Nr.", "count")
        ).reset_index()

        # Sujungimas: dokumentų kiekis + klaidų kiekis
        siuntejai_stats = siuntejai_visi.merge(siuntejai_klaidos, on="Siuntėjas", how="left")
        siuntejai_stats["Klaidų skaičius"] = siuntejai_stats["Klaidų skaičius"].fillna(0).astype(int)
        siuntejai_stats["Klaidų_procentas"] = (
            siuntejai_stats["Klaidų skaičius"] / siuntejai_stats["Dokumentų_skaičius"] * 100
        ).round(2)

        siuntejai_stats = siuntejai_stats.sort_values(
            by=["Klaidų skaičius", "Dokumentų_skaičius"],
            ascending=[False, False]
        ).reset_index(drop=True)

        # Užsakovų analizė
        uzsakovai_klaidos = klaidos["Užsakovas"].fillna("Nenurodyta").value_counts().reset_index()
        uzsakovai_klaidos.columns = ["Užsakovas", "Klaidų skaičius"]

        uzsakovai_visi = df_filtered.groupby("Užsakovas").agg(
            Dokumentų_skaičius=("Sąskaitos faktūros Nr.", "count")
        ).reset_index()

        uzsakovai_stats = uzsakovai_visi.merge(uzsakovai_klaidos, on="Užsakovas", how="left")
        uzsakovai_stats["Klaidų skaičius"] = uzsakovai_stats["Klaidų skaičius"].fillna(0).astype(int)
        uzsakovai_stats["Klaidų_procentas"] = (
            uzsakovai_stats["Klaidų skaičius"] / uzsakovai_stats["Dokumentų_skaičius"] * 100
        ).round(2)

        uzsakovai_stats = uzsakovai_stats.sort_values(
            by=["Klaidų skaičius", "Dokumentų_skaičius"],
            ascending=[False, False]
        ).reset_index(drop=True)

        # ----------------------------
        # 2 EILĖ: PRIEŽASTYS + TOP SIUNTĖJAI PAGAL KLAIDŲ KIEKĮ
        # ----------------------------
        c1, c2 = st.columns(2)

        with c1:
            st.subheader("📌 Klaidos pagal priežastį")
            st.dataframe(priezastys, use_container_width=True, height=320)

            fig_reason, ax_reason = plt.subplots(figsize=(8, 5))
            ax_reason.barh(priezastys["Klaidos priežastis"], priezastys["Klaidų skaičius"])
            ax_reason.set_title("Klaidos pagal priežastį")
            ax_reason.set_xlabel("Klaidų skaičius")
            ax_reason.invert_yaxis()
            ax_reason.grid(axis="x", alpha=0.3)
            st.pyplot(fig_reason)

        with c2:
            st.subheader("📨 TOP siuntėjai pagal klaidų kiekį")
            st.dataframe(siuntejai_stats, use_container_width=True, height=320)

            top_siuntejai_count = siuntejai_stats.head(10).copy()
            top_siuntejai_count["Siuntėjas_short"] = top_siuntejai_count["Siuntėjas"].apply(lambda x: ellipsis_label(x, 32))

            fig_sender_count, ax_sender_count = plt.subplots(figsize=(8, 5))
            ax_sender_count.barh(top_siuntejai_count["Siuntėjas_short"], top_siuntejai_count["Klaidų skaičius"])
            ax_sender_count.set_title("TOP siuntėjai pagal klaidų kiekį")
            ax_sender_count.set_xlabel("Klaidų skaičius")
            ax_sender_count.invert_yaxis()
            ax_sender_count.grid(axis="x", alpha=0.3)
            st.pyplot(fig_sender_count)

        # ----------------------------
        # 3 EILĖ: TOP SIUNTĖJAI PAGAL KLAIDŲ PROCENTĄ + TOP UŽSAKOVAI
        # ----------------------------
        c3, c4 = st.columns(2)

        with c3:
            st.subheader("📊 Siuntėjų kokybė pagal klaidų procentą")

            siuntejai_proc = siuntejai_stats[siuntejai_stats["Dokumentų_skaičius"] >= 3].copy()
            siuntejai_proc = siuntejai_proc.sort_values(
                by=["Klaidų_procentas", "Klaidų skaičius"],
                ascending=[False, False]
            ).reset_index(drop=True)

            st.caption("Rodomi siuntėjai, kurie turi bent 3 dokumentus, kad procentas nebūtų klaidinantis.")
            st.dataframe(siuntejai_proc, use_container_width=True, height=320)

            top_siuntejai_proc = siuntejai_proc.head(10).copy()
            top_siuntejai_proc["Siuntėjas_short"] = top_siuntejai_proc["Siuntėjas"].apply(lambda x: ellipsis_label(x, 32))

            fig_sender_proc, ax_sender_proc = plt.subplots(figsize=(8, 5))
            ax_sender_proc.barh(top_siuntejai_proc["Siuntėjas_short"], top_siuntejai_proc["Klaidų_procentas"])
            ax_sender_proc.set_title("TOP siuntėjai pagal klaidų procentą")
            ax_sender_proc.set_xlabel("Klaidų procentas (%)")
            ax_sender_proc.invert_yaxis()
            ax_sender_proc.grid(axis="x", alpha=0.3)
            st.pyplot(fig_sender_proc)

        with c4:
            st.subheader("🏢 TOP užsakovai su klaidomis")
            st.dataframe(uzsakovai_stats, use_container_width=True, height=320)

            top_uzsakovai = uzsakovai_stats.head(10).copy()
            top_uzsakovai["Užsakovas_short"] = top_uzsakovai["Užsakovas"].apply(lambda x: ellipsis_label(x, 32))

            fig_customer, ax_customer = plt.subplots(figsize=(8, 5))
            ax_customer.barh(top_uzsakovai["Užsakovas_short"], top_uzsakovai["Klaidų skaičius"])
            ax_customer.set_title("TOP užsakovai pagal klaidų kiekį")
            ax_customer.set_xlabel("Klaidų skaičius")
            ax_customer.invert_yaxis()
            ax_customer.grid(axis="x", alpha=0.3)
            st.pyplot(fig_customer)

        # ----------------------------
        # 4 EILĖ: PARETO
        # ----------------------------
        st.subheader("📊 Pareto analizė pagal siuntėją")

        pareto = siuntejai_stats.sort_values(by="Klaidų skaičius", ascending=False).copy()
        pareto["Kumuliacinis %"] = (
            pareto["Klaidų skaičius"].cumsum() / pareto["Klaidų skaičius"].sum() * 100
        ).round(2)

        pareto_top = pareto.head(10).copy()
        pareto_top["Siuntėjas_short"] = pareto_top["Siuntėjas"].apply(lambda x: wrap_label(x, 16))

        fig_pareto, ax1 = plt.subplots(figsize=(10, 5.8))
        ax1.bar(pareto_top["Siuntėjas_short"], pareto_top["Klaidų skaičius"])
        ax1.set_ylabel("Klaidų skaičius")
        ax1.set_xlabel("Siuntėjas")
        ax1.grid(axis="y", alpha=0.3)

        ax2 = ax1.twinx()
        ax2.plot(
            pareto_top["Siuntėjas_short"],
            pareto_top["Kumuliacinis %"],
            color="red",
            marker="o",
            linewidth=2
        )
        ax2.set_ylabel("Kumuliacinis %")
        ax2.set_ylim(0, 110)

        plt.title("Pareto analizė – TOP siuntėjai")
        plt.xticks(rotation=0, ha="center")
        st.pyplot(fig_pareto)

    else:
        priezastys = pd.DataFrame(columns=["Klaidos priežastis", "Klaidų skaičius"])
        siuntejai_stats = pd.DataFrame(columns=["Siuntėjas", "Dokumentų_skaičius", "Klaidų skaičius", "Klaidų_procentas"])
        uzsakovai_stats = pd.DataFrame(columns=["Užsakovas", "Dokumentų_skaičius", "Klaidų skaičius", "Klaidų_procentas"])
        pareto = pd.DataFrame(columns=["Siuntėjas", "Dokumentų_skaičius", "Klaidų skaičius", "Klaidų_procentas", "Kumuliacinis %"])
        st.info("Pagal pasirinktus filtrus klaidų nėra.")

    # ----------------------------
    # KLAIDŲ SĄRAŠAS
    # ----------------------------
    st.subheader("📝 Klaidų sąrašas")
    st.dataframe(klaidos.reset_index(drop=True), use_container_width=True, height=420)

    # ----------------------------
    # EXCEL EKSPORTAS
    # ----------------------------
    img_months_path = "grafikas_menesiai.png"
    img_reason_path = "grafikas_priezastys.png"
    img_sender_count_path = "grafikas_siuntejai_kiekis.png"
    img_sender_proc_path = "grafikas_siuntejai_proc.png"
    img_customer_path = "grafikas_uzsakovai.png"
    img_pareto_path = "grafikas_pareto.png"

    fig_months.savefig(img_months_path, bbox_inches="tight")

    if not klaidos.empty:
        fig_reason.savefig(img_reason_path, bbox_inches="tight")
        fig_sender_count.savefig(img_sender_count_path, bbox_inches="tight")
        fig_sender_proc.savefig(img_sender_proc_path, bbox_inches="tight")
        fig_customer.savefig(img_customer_path, bbox_inches="tight")
        fig_pareto.savefig(img_pareto_path, bbox_inches="tight")

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Suvestinė"
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws1.append(r)
    safe_add_image(ws1, img_months_path, "I2")

    ws2 = wb.create_sheet(title="Klaidų sąrašas")
    for r in dataframe_to_rows(klaidos, index=False, header=True):
        ws2.append(r)

    ws3 = wb.create_sheet(title="Priežastys")
    for r in dataframe_to_rows(priezastys, index=False, header=True):
        ws3.append(r)
    safe_add_image(ws3, img_reason_path, "D2")

    ws4 = wb.create_sheet(title="Siuntėjai")
    for r in dataframe_to_rows(siuntejai_stats, index=False, header=True):
        ws4.append(r)
    safe_add_image(ws4, img_sender_count_path, "F2")

    ws5 = wb.create_sheet(title="Siuntėjų %")
    if not siuntejai_stats.empty:
        siuntejai_proc_export = siuntejai_stats[siuntejai_stats["Dokumentų_skaičius"] >= 3].sort_values(
            by=["Klaidų_procentas", "Klaidų skaičius"],
            ascending=[False, False]
        )
    else:
        siuntejai_proc_export = pd.DataFrame(columns=siuntejai_stats.columns)

    for r in dataframe_to_rows(siuntejai_proc_export, index=False, header=True):
        ws5.append(r)
    safe_add_image(ws5, img_sender_proc_path, "F2")

    ws6 = wb.create_sheet(title="Užsakovai")
    for r in dataframe_to_rows(uzsakovai_stats, index=False, header=True):
        ws6.append(r)
    safe_add_image(ws6, img_customer_path, "F2")

    ws7 = wb.create_sheet(title="Pareto")
    for r in dataframe_to_rows(pareto, index=False, header=True):
        ws7.append(r)
    safe_add_image(ws7, img_pareto_path, "F2")

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    st.download_button(
        label="📥 Atsisiųsti Excel ataskaitą su grafikais",
        data=excel_buffer,
        file_name="Klaidu_Ataskaita.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ----------------------------
    # AI ANALIZĖ
    # ----------------------------
    st.subheader("🤖 Dirbtinio intelekto analizė")

    try:
        summary_md = summary.to_markdown(index=False)
        priezastys_md = priezastys.to_markdown(index=False) if not priezastys.empty else "Nėra klaidų."
        siuntejai_md = siuntejai_stats.to_markdown(index=False) if not siuntejai_stats.empty else "Nėra duomenų."
        uzsakovai_md = uzsakovai_stats.to_markdown(index=False) if not uzsakovai_stats.empty else "Nėra duomenų."

        analysis_prompt = f"""
Analizuok pateiktus duomenis apie sąskaitų skaičių, klaidų procentą, klaidų priežastis, siuntėjus ir užsakovus.

Svarbu:
- Atskirai įvertink absoliutų klaidų kiekį ir klaidų procentą.
- Nevertink vien tik pagal klaidų kiekį, nes kai kurie siuntėjai gali siųsti daugiau dokumentų.
- Pabrėžk, kur problema yra apimtyje, o kur – kokybėje.

Prašau:
1. Įvertinti mėnesines tendencijas
2. Nustatyti, kur koncentruojasi daugiausia klaidų
3. Įvertinti, kurie siuntėjai siunčia daugiausia klaidų absoliučiai
4. Įvertinti, kurie siuntėjai turi didžiausią klaidų procentą
5. Įvertinti, ar matomas Pareto principas
6. Išskirti didžiausią problemą procese
7. Pateikti aiškias rekomendacijas, kaip sumažinti klaidas
8. Pabrėžti, ar problema labiau susijusi su šaltiniu, procesu ar duomenų kokybe

Mėnesių suvestinė:
{summary_md}

Klaidos pagal priežastį:
{priezastys_md}

Siuntėjų analizė:
{siuntejai_md}

Užsakovų analizė:
{uzsakovai_md}
"""

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "Tu esi patyręs verslo ir procesų analitikas."},
                {"role": "user", "content": analysis_prompt}
            ],
            temperature=0.4
        )

        st.markdown(response.choices[0].message.content)

    except Exception as e:
        st.warning("Nepavyko gauti AI analizės. Patikrink API raktą Streamlit `secrets` nustatymuose.")
        st.error(str(e))
